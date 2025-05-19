import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# Create a workbook with multiple sheets
wb = Workbook()

# Define common styling
header_fill = PatternFill(start_color="2B4B8C", end_color="2B4B8C", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="6B5CA5", end_color="6B5CA5", fill_type="solid")
subheader_font = Font(bold=True, color="FFFFFF", size=11)
normal_font = Font(size=11)
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# SHEET 1: Overview and Methodology
overview_sheet = wb.active
overview_sheet.title = "1. Overview"

# Add title and headers
overview_sheet.merge_cells('A1:F1')
overview_sheet['A1'] = "Educational Framework Cost-Benefit Analysis Model"
overview_sheet['A1'].font = Font(bold=True, size=16)
overview_sheet['A1'].alignment = Alignment(horizontal='center')

# Add overview content
overview_sheet['A3'] = "Purpose:"
overview_sheet['A3'].font = Font(bold=True)
overview_sheet['A4'] = "This model quantifies the economic and social returns of adopting the Perfected Enhanced Educational Systems Implementation Framework."

overview_sheet['A6'] = "Methodology:"
overview_sheet['A6'].font = Font(bold=True)
overview_data = [
    ["Aspect", "Description"],
    ["Costs", "Direct and indirect expenses estimated per tier"],
    ["Benefits", "Financial and social benefits quantified using case model data"],
    ["Time Horizon", "5–10 years, with short-term and long-term projections"],
    ["Equity Focus", "Disaggregates benefits by marginalized groups"],
]

for row_idx, row in enumerate(overview_data, start=7):
    for col_idx, value in enumerate(row, start=1):
        cell = overview_sheet.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 7:  # Header row
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border_thin
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='left')

# Auto-fit columns
for column in overview_sheet.columns:
    max_length = 0
    try:
        column_letter = column[0].column_letter
    except AttributeError:
        # Get column letter from cell coordinate
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(column[0].column)
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = min(max_length + 2, 50)
    overview_sheet.column_dimensions[column_letter].width = adjusted_width

# SHEET 2: Cost Inputs
cost_sheet = wb.create_sheet("2. Cost Inputs")

# Add header
cost_sheet.merge_cells('A1:G1')
cost_sheet['A1'] = "Cost Estimates by Implementation Tier"
cost_sheet['A1'].font = Font(bold=True, size=14)
cost_sheet['A1'].alignment = Alignment(horizontal='center')

# Cost categories data
cost_data = [
    ["Cost Category", "Description", "Tier 1: Micro-Pilot (10-100 learners)", "Tier 2: Regional (1,000-10,000 learners)", "Tier 3: National (100,000+ learners)", "Unit Cost", "Equity Notes"],
    ["Educator Training", "40-hour facilitation training", "$10,000", "$100,000", "$1,000,000", "$1,000/educator", "Prioritize women, neurodiverse trainers"],
    ["Curriculum Materials", "Spiral dynamics modules, regenerative guides", "$5,000", "$50,000", "$500,000", "Varies", "Free for low-income regions"],
    ["Technology", "Digital platforms, low-tech alternatives", "$10,000", "$100,000", "$1,000,000", "Varies", "Subsidized for rural areas"],
    ["Community Engagement", "Workshops, forums", "$5,000", "$50,000", "$500,000", "Per event", "50% marginalized representation"],
    ["M&E", "Rubrics, data collection", "$5,000", "$50,000", "$500,000", "Per metric", "Anonymous data for safety"],
]

# Add cost data
for row_idx, row in enumerate(cost_data, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = cost_sheet.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 3:  # Header row
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border_thin
        if col_idx in [3, 4, 5] and row_idx > 3:  # Format currency columns
            try:
                # Remove $ and commas for numeric formatting
                numeric_value = float(value.replace('$', '').replace(',', ''))
                cell.value = numeric_value
                cell.number_format = '#,##0'
            except:
                pass

# Add total rows
total_row = cost_sheet.max_row + 1
cost_sheet.cell(row=total_row, column=1, value="Total Estimated Costs:")
cost_sheet.cell(row=total_row, column=1).font = Font(bold=True)

# Calculate totals for each tier
for col in range(3, 6):
    col_letter = chr(ord('A') + col - 1)
    sum_formula = f"=SUM({col_letter}4:{col_letter}{total_row-1})"
    total_cell = cost_sheet.cell(row=total_row, column=col, value=sum_formula)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '$#,##0'

# Add customization note
note_row = total_row + 2
cost_sheet.cell(row=note_row, column=1, value="Note:")
cost_sheet.cell(row=note_row, column=1).font = Font(bold=True, color="FF0000")
cost_sheet.cell(row=note_row, column=2, value="Adjust costs based on local wages and resource availability")

# Auto-fit columns
for column in cost_sheet.columns:
    max_length = 0
    try:
        column_letter = column[0].column_letter
    except AttributeError:
        # Get column letter from cell coordinate
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(column[0].column)
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = min(max_length + 2, 50)
    cost_sheet.column_dimensions[column_letter].width = adjusted_width

# SHEET 3: Benefit Projections
benefit_sheet = wb.create_sheet("3. Benefit Projections")

# Add header
benefit_sheet.merge_cells('A1:G1')
benefit_sheet['A1'] = "Benefit Projections by Implementation Tier"
benefit_sheet['A1'].font = Font(bold=True, size=14)
benefit_sheet['A1'].alignment = Alignment(horizontal='center')

# Benefit categories data
benefit_data = [
    ["Benefit Category", "Description", "Tier 1: Micro-Pilot", "Tier 2: Regional", "Tier 3: National", "Unit Measure", "Equity Notes"],
    ["Economic Returns", "Increased literacy, job creation", "$1,000,000", "$10,000,000", "$2,000,000,000", "Over 5-10 years", "Prioritize jobs for refugees"],
    ["Educational Outcomes", "Systems thinking, empathy gains", "80% proficiency", "85% proficiency", "90% proficiency", "Percentage", "Oral assessments for neurodiverse"],
    ["Social Equity", "Diversity in governance", "90% equity index", "95% equity index", "98% equity index", "Index score", "30% marginalized representation"],
    ["Environmental Impact", "Regenerative projects", "10 hectares", "100 hectares", "1,000 hectares", "Hectares restored", "Indigenous-led restoration"],
    ["Civic Engagement", "Youth-led policies", "5 policies", "50 policies", "500 policies", "Policies proposed", "LGBTQ+, refugee voices prioritized"],
]

# Add benefit data
for row_idx, row in enumerate(benefit_data, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = benefit_sheet.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 3:  # Header row
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border_thin
        # Only format economic returns as currency
        if col_idx in [3, 4, 5] and row_idx == 4:
            try:
                numeric_value = float(value.replace('$', '').replace(',', ''))
                cell.value = numeric_value
                cell.number_format = '#,##0'
            except:
                pass

# Auto-fit columns
for column in benefit_sheet.columns:
    max_length = 0
    try:
        column_letter = column[0].column_letter
    except AttributeError:
        # Get column letter from cell coordinate
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(column[0].column)
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = min(max_length + 2, 50)
    benefit_sheet.column_dimensions[column_letter].width = adjusted_width

# SHEET 4: Analysis and Calculations
analysis_sheet = wb.create_sheet("4. Analysis")

# Add header
analysis_sheet.merge_cells('A1:E1')
analysis_sheet['A1'] = "ROI Analysis by Implementation Tier"
analysis_sheet['A1'].font = Font(bold=True, size=14)
analysis_sheet['A1'].alignment = Alignment(horizontal='center')

# Analysis headers
analysis_headers = ["Implementation Tier", "Total Costs", "Total Benefits", "Net Benefit", "ROI (%)"]
for col_idx, header in enumerate(analysis_headers, start=1):
    cell = analysis_sheet.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border_thin

# Implementation tiers
tiers = ["Tier 1: Micro-Pilot", "Tier 2: Regional", "Tier 3: National"]
for row_idx, tier in enumerate(tiers, start=4):
    # Tier name
    tier_cell = analysis_sheet.cell(row=row_idx, column=1, value=tier)
    tier_cell.font = Font(bold=True)
    tier_cell.border = border_thin
    
    # Cost reference
    cost_cell = analysis_sheet.cell(row=row_idx, column=2, value=f"='2. Cost Inputs'!{chr(ord('C') + row_idx - 4)}9")
    cost_cell.number_format = '$#,##0'
    cost_cell.border = border_thin
    
    # Benefit reference (economic returns only)
    benefit_cell = analysis_sheet.cell(row=row_idx, column=3, value=f"='3. Benefit Projections'!{chr(ord('C') + row_idx - 4)}4")
    benefit_cell.number_format = '$#,##0'
    benefit_cell.border = border_thin
    
    # Net benefit calculation
    net_benefit_cell = analysis_sheet.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}")
    net_benefit_cell.number_format = '$#,##0'
    net_benefit_cell.border = border_thin
    net_benefit_cell.font = Font(bold=True)
    
    # ROI calculation
    roi_cell = analysis_sheet.cell(row=row_idx, column=5, value=f"=(D{row_idx}/B{row_idx})*100")
    roi_cell.number_format = '0.0%'
    roi_cell.border = border_thin
    roi_cell.font = Font(bold=True, color="00FF00")

# Add interpretation section
analysis_sheet.cell(row=9, column=1, value="Interpretation:")
analysis_sheet.cell(row=9, column=1).font = Font(bold=True)
analysis_sheet.cell(row=10, column=1, value="• ROI represents return on investment as a percentage")
analysis_sheet.cell(row=11, column=1, value="• Net benefit shows absolute financial gain after costs")
analysis_sheet.cell(row=12, column=1, value="• Higher tiers show greater total impact and efficiency")

# Auto-fit columns
for column in analysis_sheet.columns:
    max_length = 0
    try:
        column_letter = column[0].column_letter
    except AttributeError:
        # Get column letter from cell coordinate
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(column[0].column)
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = min(max_length + 2, 50)
    analysis_sheet.column_dimensions[column_letter].width = adjusted_width

# SHEET 5: Summary Dashboard
dashboard_sheet = wb.create_sheet("5. Dashboard")

# Add header
dashboard_sheet.merge_cells('A1:G1')
dashboard_sheet['A1'] = "Educational Framework Cost-Benefit Analysis Dashboard"
dashboard_sheet['A1'].font = Font(bold=True, size=16)
dashboard_sheet['A1'].alignment = Alignment(horizontal='center')

# Key metrics summary
dashboard_sheet.cell(row=3, column=1, value="Key Financial Metrics:")
dashboard_sheet.cell(row=3, column=1).font = Font(bold=True, size=14)

# Summary table
summary_headers = ["Metric", "Tier 1", "Tier 2", "Tier 3"]
for col_idx, header in enumerate(summary_headers, start=1):
    cell = dashboard_sheet.cell(row=5, column=col_idx, value=header)
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.border = border_thin

metrics = [
    ["Total Investment", "=IF('4. Analysis'!B4=0,\"\",TEXT('4. Analysis'!B4,\"$#,##0\"))", "=IF('4. Analysis'!B5=0,\"\",TEXT('4. Analysis'!B5,\"$#,##0\"))", "=IF('4. Analysis'!B6=0,\"\",TEXT('4. Analysis'!B6,\"$#,##0\"))"],
    ["Expected Returns", "=IF('4. Analysis'!C4=0,\"\",TEXT('4. Analysis'!C4,\"$#,##0\"))", "=IF('4. Analysis'!C5=0,\"\",TEXT('4. Analysis'!C5,\"$#,##0\"))", "=IF('4. Analysis'!C6=0,\"\",TEXT('4. Analysis'!C6,\"$#,##0\"))"],
    ["Net Benefit", "=IF('4. Analysis'!D4=0,\"\",TEXT('4. Analysis'!D4,\"$#,##0\"))", "=IF('4. Analysis'!D5=0,\"\",TEXT('4. Analysis'!D5,\"$#,##0\"))", "=IF('4. Analysis'!D6=0,\"\",TEXT('4. Analysis'!D6,\"$#,##0\"))"],
    ["ROI", "=IF('4. Analysis'!E4=0,\"\",TEXT('4. Analysis'!E4,\"0.0%\"))", "=IF('4. Analysis'!E5=0,\"\",TEXT('4. Analysis'!E5,\"0.0%\"))", "=IF('4. Analysis'!E6=0,\"\",TEXT('4. Analysis'!E6,\"0.0%\"))"],
]

for row_idx, metric in enumerate(metrics, start=6):
    for col_idx, value in enumerate(metric, start=1):
        cell = dashboard_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border_thin
        if col_idx == 1:
            cell.font = Font(bold=True)
        else:
            cell.alignment = Alignment(horizontal='center')

# Add social impact metrics
dashboard_sheet.cell(row=12, column=1, value="Social Impact Indicators:")
dashboard_sheet.cell(row=12, column=1).font = Font(bold=True, size=14)

social_metrics = [
    ["Systems Thinking Proficiency", "80%", "85%", "90%"],
    ["Equity Index", "90%", "95%", "98%"],
    ["Hectares Restored", "10", "100", "1,000"],
    ["Youth Policies Initiated", "5", "50", "500"],
]

# Social metrics headers
for col_idx, header in enumerate(summary_headers, start=1):
    cell = dashboard_sheet.cell(row=14, column=col_idx, value=header)
    cell.fill = PatternFill(start_color="2D5F2D", end_color="2D5F2D", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border_thin

for row_idx, metric in enumerate(social_metrics, start=15):
    for col_idx, value in enumerate(metric, start=1):
        cell = dashboard_sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border_thin
        if col_idx == 1:
            cell.font = Font(bold=True)
        else:
            cell.alignment = Alignment(horizontal='center')

# Add notes
dashboard_sheet.cell(row=20, column=1, value="Instructions:")
dashboard_sheet.cell(row=20, column=1).font = Font(bold=True, color="FF0000")
instructions = [
    "1. Customize costs in Sheet 2 based on local wages and resources",
    "2. Adjust benefits in Sheet 3 using local pilot data",
    "3. Use Sheet 4 for ROI calculations and advocacy materials",
    "4. This dashboard provides an at-a-glance view of all metrics",
]
for row_idx, instruction in enumerate(instructions, start=21):
    dashboard_sheet.cell(row=row_idx, column=1, value=instruction)

# Auto-fit columns
for column in dashboard_sheet.columns:
    max_length = 0
    try:
        column_letter = column[0].column_letter
    except AttributeError:
        # Get column letter from cell coordinate
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(column[0].column)
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = min(max_length + 2, 50)
    dashboard_sheet.column_dimensions[column_letter].width = adjusted_width

# SHEET 6: Instructions
instructions_sheet = wb.create_sheet("6. Instructions")

# Add header
instructions_sheet.merge_cells('A1:E1')
instructions_sheet['A1'] = "How to Use This Cost-Benefit Analysis Model"
instructions_sheet['A1'].font = Font(bold=True, size=16)
instructions_sheet['A1'].alignment = Alignment(horizontal='center')

# Instructions content
instructions_content = [
    ["Step", "Action", "Details"],
    ["1", "Gather Data", "Collect local cost data (wages, materials) and benefit projections"],
    ["2", "Customize Model", "Input data into Sheet 2 (Costs) and Sheet 3 (Benefits)"],
    ["3", "Validate", "Engage community boards with 50% marginalized representation"],
    ["4", "Analyze", "Review Sheet 4 (Analysis) for ROI calculations"],
    ["5", "Create Materials", "Use Sheet 5 (Dashboard) for presentations and advocacy"],
    ["6", "Share & Iterate", "Distribute results and refine based on feedback"],
]

for row_idx, row in enumerate(instructions_content, start=3):
    for col_idx, value in enumerate(row, start=1):
        cell = instructions_sheet.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 3:  # Header row
            cell.fill = header_fill
            cell.font = header_font
        elif col_idx == 1 and row_idx > 3:  # Step numbers
            cell.font = Font(bold=True)
        cell.border = border_thin

# Add tips
tips_row = instructions_sheet.max_row + 2
instructions_sheet.cell(row=tips_row, column=1, value="Tips:")
instructions_sheet.cell(row=tips_row, column=1).font = Font(bold=True)

tips = [
    "• Validate assumptions with local stakeholders",
    "• Consider both financial and social returns",
    "• Prioritize equity in all calculations",
    "• Adjust time horizons based on context",
    "• Document all customizations for transparency",
]

for row_idx, tip in enumerate(tips, start=tips_row+1):
    instructions_sheet.cell(row=row_idx, column=1, value=tip)

# Auto-fit columns
for column in instructions_sheet.columns:
    max_length = 0
    try:
        column_letter = column[0].column_letter
    except AttributeError:
        # Get column letter from cell coordinate
        from openpyxl.utils import get_column_letter
        column_letter = get_column_letter(column[0].column)
    for cell in column:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    adjusted_width = min(max_length + 2, 50)
    instructions_sheet.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
wb.save('cost-benefit-analysis-model-en.xlsx')
print("Excel file created successfully!")
